"""
Fuel Price Tracker v7 — refactor
Changes vs v6.1:
  1. Orlen LT: parse by COLUMN POSITION (Pardavimo kaina su PVM = 5th col),
     require "Dyzelinas ... su RRME", EXCLUDE agri/marine/heating, explicit
     terminal selection. No more fragile "max number in line".
  2. ELVIS DE: real public German diesel — Tankerkönig (MTS-K) when an API key
     is set, else EC Oil Bulletin Germany diesel fallback. (Real ELVIS
     Dieselfloater is partner-only / not public.)
  3. validate_fx() + validate_price_change() implemented; outlier guard wired
     into the Excel writer (>MAX_DAILY_CHANGE_PCT → flagged SUSPECT, not silent).
  4. Centralized config.py is now actually used.
"""

import requests, json, re, io, sys, calendar
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config as cfg

EXCEL_PATH = cfg.EXCEL_PATH
H = cfg.HEADERS
TODAY = datetime.now()
TODAY_STR = TODAY.strftime("%Y-%m-%d")
WDAY = TODAY.weekday()

# Collected by update_excel(), surfaced in latest_results.json
SUSPECTS = []


def log(s, m, l="INFO"):
    print(f"[{l}] {s}: {m}")


def clean_num(s):
    """Parse numbers like '1 756.06', '6 192', '1,234.56', '1234,56'"""
    if s is None:
        return None
    s = str(s).strip()
    s = re.sub(r'[€$£\xa0]', '', s)
    if re.match(r'^\d[\d ]+\.\d+$', s):
        return float(s.replace(' ', ''))
    if re.match(r'^\d[\d ]+,\d+$', s):
        return float(s.replace(' ', '').replace(',', '.'))
    if ',' in s and '.' in s:
        return float(s.replace(',', ''))
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    s = s.replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return None


# ═══════════════════════════════════════
# VALIDATION (used by tests + outlier guard)
# ═══════════════════════════════════════
def validate_fx(pln_eur, sek_eur):
    """True if both FX rates are within sane bounds (inclusive)."""
    if pln_eur is None or sek_eur is None:
        return False
    return (cfg.FX_PLN_EUR_MIN <= pln_eur <= cfg.FX_PLN_EUR_MAX and
            cfg.FX_SEK_EUR_MIN <= sek_eur <= cfg.FX_SEK_EUR_MAX)


def validate_price_change(name, new_val, prev_val, max_pct=None):
    """
    True if the day-over-day change is acceptable (< max_pct).
    No previous value (None / 0) → True. A change >= max_pct → False.
    """
    if max_pct is None:
        max_pct = cfg.MAX_DAILY_CHANGE_PCT
    if new_val is None:
        return False
    if not prev_val:  # None or 0
        return True
    change_pct = abs((new_val - prev_val) / prev_val) * 100
    ok = change_pct < max_pct
    if not ok:
        log(name, f"OUTLIER: {prev_val} → {new_val} ({change_pct:.1f}% ≥ {max_pct}%)", "WARN")
    return ok


# ═══════════════════════════════════════
# 1. FX RATES — with multiple fallback APIs
# ═══════════════════════════════════════
def fetch_fx():
    import time
    try:
        r = requests.get(cfg.URLS["fx_frankfurter"], timeout=cfg.FX_TIMEOUT)
        r.raise_for_status()
        d = r.json().get("rates", {})
        if d.get("PLN") and d.get("SEK"):
            log("FX", f"frankfurter.app: PLN={d['PLN']}, SEK={d['SEK']}")
            return {"PLN_EUR": d["PLN"], "SEK_EUR": d["SEK"]}
    except Exception as e:
        log("FX", f"frankfurter.app failed: {e}", "WARN")
        time.sleep(2)
    try:
        r = requests.get(cfg.URLS["fx_exchangerate"], timeout=cfg.FX_TIMEOUT)
        r.raise_for_status()
        d = r.json().get("rates", {})
        if d.get("PLN") and d.get("SEK"):
            log("FX", f"exchangerate.host: PLN={d['PLN']}, SEK={d['SEK']}")
            return {"PLN_EUR": d["PLN"], "SEK_EUR": d["SEK"]}
    except Exception as e:
        log("FX", f"exchangerate.host failed: {e}", "WARN")
        time.sleep(2)
    try:
        r = requests.get(cfg.URLS["fx_ecb_xml"], timeout=cfg.FX_TIMEOUT)
        r.raise_for_status()
        pln = re.search(r"currency='PLN'\s+rate='([\d.]+)'", r.text)
        sek = re.search(r"currency='SEK'\s+rate='([\d.]+)'", r.text)
        if pln and sek:
            pln_v, sek_v = float(pln.group(1)), float(sek.group(1))
            log("FX", f"ECB XML: PLN={pln_v}, SEK={sek_v}")
            return {"PLN_EUR": pln_v, "SEK_EUR": sek_v}
    except Exception as e:
        log("FX", f"ECB XML failed: {e}", "WARN")
    log("FX", "All 3 sources failed!", "ERROR")
    return None


# ═══════════════════════════════════════
# 2. ORLEN PL — via petrodom.pl
# ═══════════════════════════════════════
def fetch_orlen_pl():
    import time
    urls = [cfg.URLS["orlen_pl"], cfg.URLS["orlen_pl_alt"]]
    for attempt, url in enumerate(urls):
        try:
            log("Orlen PL", f"Trying URL {attempt+1}/{len(urls)}")
            r = requests.get(url, headers=H, timeout=cfg.REQUEST_TIMEOUT)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for table in soup.find_all("table"):
                for row in table.find_all("tr"):
                    cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                    for i, cell in enumerate(cells):
                        if "ekodiesel" in cell.lower() and "arktyczny" not in cell.lower() and "grzewczy" not in cell.lower():
                            for j in range(i + 1, min(i + 3, len(cells))):
                                price = clean_num(cells[j])
                                if price and cfg.ORLEN_PL_MIN < price < cfg.ORLEN_PL_MAX:
                                    log("Orlen PL", f"Ekodiesel = {price} PLN/m³")
                                    return {"price_pln_m3": price}
            text = soup.get_text(" ", strip=True)
            m = re.search(r'[Ee]kodiesel[^0-9]{0,60}?(\d[\d\s\xa0]*\d)', text)
            if m:
                price = clean_num(m.group(1))
                if price and cfg.ORLEN_PL_MIN < price < cfg.ORLEN_PL_MAX:
                    log("Orlen PL", f"Ekodiesel (text fallback) = {price} PLN/m³")
                    return {"price_pln_m3": price}
            log("Orlen PL", f"URL {attempt+1} — Ekodiesel not found", "WARN")
            time.sleep(2)
        except Exception as e:
            log("Orlen PL", f"URL {attempt+1} error: {e}", "WARN")
            time.sleep(2)
    log("Orlen PL", "All URLs failed", "ERROR")
    return None


# ═══════════════════════════════════════
# 3. ORLEN LT — PDF: parse by COLUMN
# ═══════════════════════════════════════
def get_season_classes(month=None):
    """EN 590 diesel class letters expected for the given month."""
    if month is None:
        month = TODAY.month
    for months, classes in cfg.ORLEN_LT_SEASON_CLASSES.items():
        if month in months:
            return classes
    return ["C", "E"]


def _line_is_target_diesel(line):
    """Road diesel 'Dyzelinas ... su RRME', not agri/marine/heating."""
    low = line.lower()
    if not all(req.lower() in low for req in cfg.ORLEN_LT_REQUIRE):
        return False
    if any(ex.lower() in low for ex in cfg.ORLEN_LT_EXCLUDE):
        return False
    return True


def _nums_in_line(line):
    """All EUR/1000l style numbers in order: '836.77', '1 340.37', ..."""
    return [float(n.replace(' ', '')) for n in re.findall(r'(\d[\d ]*\.\d{2,5})', line)]


def _find_orlen_lt_prices(text):
    """
    Parse the Orlen LT realizacija PDF text. Columns per product line:
      0 Bazinė | 1 Akcizas | 2 Bazė+akcizas(be PVM) | 3 PVM | 4 Pardavimo su PVM
    Returns dict with the configured price column (default = gross/su PVM),
    using the configured terminal (default = first listed), or the average.
    """
    classes = get_season_classes()
    m_date = re.search(r'galioja nuo\s+(\d{4}-\d{2}-\d{2})', text)
    pdf_date = m_date.group(1) if m_date else None
    candidates = []  # list of dicts per terminal occurrence
    for line in text.split('\n'):
        if not _line_is_target_diesel(line):
            continue
        nums = _nums_in_line(line)
        if len(nums) < 5:
            log("Orlen LT", f"Skip (only {len(nums)} cols): {line[:80]}", "WARN")
            continue
        cols = nums[:5]
        gross, net = cols[cfg.ORLEN_LT_COL_GROSS], cols[cfg.ORLEN_LT_COL_NET]
        if not (cfg.ORLEN_LT_MIN < gross < cfg.ORLEN_LT_MAX):
            continue
        # detect class letter present in the line (e.g. "C kl")
        cls = next((c for c in ["0", "1", "2", "C", "E"] if re.search(rf'\b{c}\s*kl', line)), "?")
        candidates.append({"gross": gross, "net": net,
                           "price": cols[cfg.ORLEN_LT_PRICE_COL], "class": cls, "line": line})

    if not candidates:
        log("Orlen LT", "No 'Dyzelinas...su RRME' line with 5 columns found", "WARN")
        return None

    # Prefer the seasonal class if several classes are present
    seasonal = [c for c in candidates if c["class"] in classes]
    chosen_set = seasonal if seasonal else candidates
    log("Orlen LT", f"{len(candidates)} terminal line(s); season={classes}; using {len(chosen_set)}")

    if cfg.ORLEN_LT_TERMINAL_INDEX is None:
        n = len(chosen_set)
        price = round(sum(c["price"] for c in chosen_set) / n, 2)
        gross = round(sum(c["gross"] for c in chosen_set) / n, 2)
        net = round(sum(c["net"] for c in chosen_set) / n, 2)
        cls = chosen_set[0]["class"]
        log("Orlen LT", f"AVG over {n} terminals: price={price}")
    else:
        idx = min(cfg.ORLEN_LT_TERMINAL_INDEX, len(chosen_set) - 1)
        c = chosen_set[idx]
        price, gross, net, cls = c["price"], c["gross"], c["net"], c["class"]
        log("Orlen LT", f"Terminal[{idx}] {c['line'][:70]}")

    eur_l = round(price / 1000, 4)
    log("Orlen LT", f"Selected: {price} EUR/1000l → {eur_l} EUR/l (su PVM={gross}, be PVM={net}, kl={cls})")
    return {
        "price_eur_l": eur_l,
        "price_eur_1000l_su_pvm": gross,
        "price_eur_1000l_be_pvm": net,
        "diesel_class": f"Dyzelinas {cls} kl",
        "pdf_date": pdf_date,
    }


def parse_orlen_lt_pdf(pdf_bytes):
    try:
        import pdfplumber
    except ImportError:
        log("Orlen LT", "pdfplumber not installed", "WARN")
        return None
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ""
        return _find_orlen_lt_prices(text)
    except Exception as e:
        log("Orlen LT", f"PDF error: {e}", "WARN")
        return None


def _orlen_lt_url_for(d):
    return f"{cfg.URLS['orlen_lt_base']}/LT/Wholesale/Prices/Kainos {d.strftime('%Y %m %d')} realizacija internet.pdf"


def fetch_orlen_lt():
    """
    Prefer the PDF whose internal validity date == today. If only an older PDF
    is available (Orlen often publishes late), return the newest one tagged with
    its real `pdf_date` so update_excel() writes it to the CORRECT date row —
    this fixes the historical off-by-one (each day showed the prior day's price).
    """
    try:
        urls = [_orlen_lt_url_for(TODAY)]  # try today's PDF first
        try:
            r = requests.get(cfg.URLS["orlen_lt_list"], headers=H, timeout=cfg.REQUEST_TIMEOUT)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if ".pdf" in href.lower() and "kainos" in href.lower():
                    if not href.startswith("http"):
                        href = cfg.URLS["orlen_lt_base"] + href
                    urls.append(href)
        except Exception as e:
            log("Orlen LT", f"listing failed: {e}", "WARN")
        for days_back in range(7):
            urls.append(_orlen_lt_url_for(TODAY - timedelta(days=days_back)))

        seen, best = set(), None
        for pdf_url in urls:
            if pdf_url in seen:
                continue
            seen.add(pdf_url)
            try:
                r2 = requests.get(pdf_url, headers=H, timeout=cfg.REQUEST_TIMEOUT)
                if r2.status_code == 200 and len(r2.content) > 500:
                    res = parse_orlen_lt_pdf(r2.content)
                    if res:
                        log("Orlen LT", f"Parsed {pdf_url.split('/')[-1]} → pdf_date={res.get('pdf_date')}")
                        if res.get("pdf_date") == TODAY_STR:
                            return res
                        if best is None:
                            best = res
            except Exception:
                continue
            if len(seen) >= 8:
                break
        if best:
            log("Orlen LT", f"Today's PDF unavailable; newest is {best.get('pdf_date')} → will route to that date", "WARN")
        else:
            log("Orlen LT", "No PDF parsed", "WARN")
        return best
    except Exception as e:
        log("Orlen LT", str(e), "ERROR")
        return None


# ═══════════════════════════════════════
# 4. ELVIS DE — German diesel (Tankerkönig → EC bulletin fallback)
# ═══════════════════════════════════════
def fetch_elvis_de():
    """
    Public German diesel reference. The real ELVIS Dieselfloater is partner-only
    (BLUE.net), so we approximate with Tankerkönig (official MTS-K pump data) if
    an API key is configured; otherwise return None and let update_excel() fall
    back to the EC Oil Bulletin Germany diesel.
    """
    if cfg.ELVIS_DE_SOURCE == "tankerkoenig" and cfg.TANKERKOENIG_API_KEY:
        city_avgs = []
        for name, lat, lng in cfg.TANKERKOENIG_CITIES:
            try:
                r = requests.get(cfg.TANKERKOENIG_URL, headers=H, timeout=cfg.REQUEST_TIMEOUT,
                                 params={"lat": lat, "lng": lng, "rad": cfg.TANKERKOENIG_RADIUS_KM,
                                         "sort": "dist", "type": "diesel",
                                         "apikey": cfg.TANKERKOENIG_API_KEY})
                r.raise_for_status()
                d = r.json()
                if not d.get("ok"):
                    log("Elvis DE", f"Tankerkönig {name}: ok=false status={d.get('status')} msg={d.get('message')}", "WARN")
                    continue
                stations = d.get("stations", [])
                if name == cfg.TANKERKOENIG_CITIES[0][0]:
                    s0 = stations[0] if stations else None
                    log("Elvis DE", f"{name}: {len(stations)} stations; keys={sorted(s0.keys()) if s0 else None}; price={s0.get('price') if s0 else None} diesel={s0.get('diesel') if s0 else None}")
                vals = []
                for s in stations:
                    p = s.get("price")
                    if isinstance(p, bool) or not isinstance(p, (int, float)):
                        p = s.get("diesel")
                    if isinstance(p, (int, float)) and not isinstance(p, bool) and p > 0:
                        vals.append(p)
                if vals:
                    city_avgs.append(sum(vals) / len(vals))
                else:
                    log("Elvis DE", f"{name}: ok but 0 usable diesel prices ({len(stations)} stations)", "WARN")
            except Exception as e:
                log("Elvis DE", f"Tankerkönig {name} failed: {e}", "WARN")
        if city_avgs:
            avg = round(sum(city_avgs) / len(city_avgs), 4)
            if cfg.DIESEL_EUR_MIN < avg < cfg.DIESEL_EUR_MAX:
                log("Elvis DE", f"Tankerkönig DE avg = {avg} EUR/l ({len(city_avgs)} cities)")
                return {"price_eur_l": avg, "source": "tankerkoenig"}
        log("Elvis DE", "Tankerkönig returned no usable data → EC bulletin fallback", "WARN")
        return None
    # No key configured → EC bulletin DE is used as the source (in update_excel)
    log("Elvis DE", "No Tankerkönig key → using EC Oil Bulletin DE diesel")
    return None


# ═══════════════════════════════════════
# 5. BSH / ST1 SE
# ═══════════════════════════════════════
def fetch_bsh_se():
    try:
        r = requests.get(cfg.URLS["bsh_se"], headers=H, timeout=cfg.REQUEST_TIMEOUT)
        r.raise_for_status()
        text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)
        for pat in [r'[Dd]iesel\s*(?:MK[123])?\s*[^0-9]{0,30}?(\d{1,2}[.,]\d{2})',
                    r'(\d{1,2}[.,]\d{2})\s*(?:kr|SEK)[^0-9]{0,20}[Dd]iesel']:
            m = re.search(pat, text)
            if m:
                p = float(m.group(1).replace(",", "."))
                if cfg.BSH_SE_MIN < p < cfg.BSH_SE_MAX:
                    log("BSH SE", f"Diesel = {p} SEK/l")
                    return {"price_sek_l": p}
        log("BSH SE", "Not found", "WARN")
        return None
    except Exception as e:
        log("BSH SE", str(e), "ERROR")
        return None


# ═══════════════════════════════════════
# 6. EU BULLETIN — direct EC XLSX download
# ═══════════════════════════════════════
def fetch_eu_bulletin():
    """Download EC Weekly Oil Bulletin XLSX — DIESEL column (Gas oil automobile)."""
    try:
        log("EU Bulletin", "Downloading EC XLSX...")
        r = requests.get(cfg.URLS["eu_bulletin_xlsx"], headers=H, timeout=cfg.FX_TIMEOUT)
        if r.status_code != 200:
            log("EU Bulletin", f"EC HTTP {r.status_code}", "WARN")
            return fetch_eu_bulletin_fallback()
        wb = load_workbook(io.BytesIO(r.content), data_only=True)
        countries = dict(cfg.EU_COUNTRIES)
        cc_names = cfg.EU_COUNTRY_NAMES
        eu_avg = de_diesel = ec_date = None
        ws = wb[wb.sheetnames[0]]
        log("EU Bulletin", f"Sheet: {ws.title} ({ws.max_row}r x {ws.max_column}c)")
        diesel_col = 3
        for row in range(1, 6):
            for col in range(1, min(10, ws.max_column + 1)):
                val = str(ws.cell(row=row, column=col).value or "").lower()
                if "gas oil" in val or "diesel" in val or "gasoil" in val:
                    diesel_col = col
                    break
            if diesel_col != 3:
                break
        log("EU Bulletin", f"Using diesel column: {diesel_col}")
        for row in range(1, 6):
            for col in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                if hasattr(val, 'strftime'):
                    ec_date = val.strftime('%Y-%m-%d')
                elif isinstance(val, str):
                    dm = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', val)
                    if dm:
                        d, m, y = dm.group(1), dm.group(2), dm.group(3)
                        ec_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"

        def conv(v):
            v = float(v)
            if 0.5 < v < 3.5:
                return round(v, 4)
            if 500 < v < 3500:
                return round(v / 1000, 4)
            return None

        for row in range(1, ws.max_row + 1):
            cell0 = str(ws.cell(row=row, column=1).value or "").strip()
            if not cell0:
                continue
            for cname, cc in cc_names.items():
                if cname.lower() in cell0.lower():
                    val = ws.cell(row=row, column=diesel_col).value
                    if val is not None:
                        try:
                            countries[cc] = conv(val)
                        except Exception:
                            pass
            if "germany" in cell0.lower():
                val = ws.cell(row=row, column=diesel_col).value
                if val is not None:
                    try:
                        de_diesel = conv(val)
                    except Exception:
                        pass
            c0l = cell0.lower()
            if eu_avg is None and ("ce/ec" in c0l or "eur27" in c0l or "eu" in c0l) and \
               ("average" in c0l or "weighted" in c0l or "moyenne" in c0l or "durchschnitt" in c0l):
                if "euro area" not in c0l and "eurozone" not in c0l:
                    val = ws.cell(row=row, column=diesel_col).value
                    if val is not None:
                        try:
                            eu_avg = conv(val)
                        except Exception:
                            pass
        found = {k: v for k, v in countries.items() if v is not None}
        if found:
            log("EU Bulletin", f"DIESEL: {found}, avg={eu_avg}, DE={de_diesel}, date={ec_date}")
            result = {**countries, "EU_AVG": eu_avg}
            if de_diesel:
                result["DE"] = de_diesel
            if ec_date:
                result["_date"] = ec_date
            return result
        log("EU Bulletin", "No diesel data found in EC XLSX", "WARN")
        return fetch_eu_bulletin_fallback()
    except Exception as e:
        log("EU Bulletin", f"EC XLSX error: {e}", "WARN")
        return fetch_eu_bulletin_fallback()


def fetch_eu_bulletin_fallback():
    """Fallback: fuel-prices.eu"""
    try:
        r = requests.get(cfg.URLS["elvis_de_fallback"], headers=H, timeout=cfg.REQUEST_TIMEOUT)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        countries = dict(cfg.EU_COUNTRIES)
        cc_map = {"lithuania": "LT", "latvia": "LV", "estonia": "EE",
                  "denmark": "DK", "sweden": "SE", "finland": "FI"}
        eu_avg = None
        for table in soup.find_all("table"):
            if "diesel" not in table.get_text(" ", strip=True).lower():
                continue
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                if len(cells) < 2:
                    continue
                rt = " ".join(cells[:2]).lower()
                for name, cc in cc_map.items():
                    if name in rt:
                        for cell in cells:
                            m = re.search(r'€?(\d\.\d{3})', cell)
                            if m:
                                val = float(m.group(1))
                                if 0.5 < val < 3.5:
                                    countries[cc] = val
                                    break
        m = re.search(r'€(\d\.\d{3})/L\s+for\s+diesel', soup.get_text(" ", strip=True))
        if m:
            eu_avg = float(m.group(1))
        found = {k: v for k, v in countries.items() if v}
        if found:
            log("EU Bulletin fallback", f"{found}, avg={eu_avg}")
            return {**countries, "EU_AVG": eu_avg}
        return None
    except Exception as e:
        log("EU Bulletin fallback", str(e), "ERROR")
        return None


# ═══════════════════════════════════════
# EXCEL helpers
# ═══════════════════════════════════════
def date_exists_in_daily(ws, target_date):
    for row in range(5, min(15, ws.max_row + 1)):
        cell = ws.cell(row=row, column=1).value
        if cell is None:
            continue
        if hasattr(cell, 'date'):
            if cell.date() == target_date.date():
                return row
        elif hasattr(cell, 'strftime'):
            if cell.strftime('%Y-%m-%d') == target_date.strftime('%Y-%m-%d'):
                return row
        elif isinstance(cell, str) and cell[:10] == target_date.strftime('%Y-%m-%d'):
            return row
    return None


def date_exists_in_weekly(ws, date_str):
    for row in range(4, min(ws.max_row + 1, 100)):
        cell = ws.cell(row=row, column=1).value
        if cell is None:
            continue
        if hasattr(cell, 'strftime'):
            if cell.strftime('%Y-%m-%d') == date_str:
                return row
        elif hasattr(cell, 'date'):
            if cell.date().strftime('%Y-%m-%d') == date_str:
                return row
        elif isinstance(cell, str) and cell[:10] == date_str:
            return row
    return None


def _prev_daily_values(ws):
    """Most recent existing daily row (date != today) — for outlier guard."""
    for row in range(5, min(ws.max_row + 1, 40)):
        c = ws.cell(row=row, column=1).value
        if c is None:
            continue
        ds = c.strftime('%Y-%m-%d') if hasattr(c, 'strftime') else str(c)[:10]
        if ds == TODAY_STR:
            continue
        return {"PL": ws.cell(row=row, column=3).value, "LT": ws.cell(row=row, column=6).value,
                "DE": ws.cell(row=row, column=9).value, "SE": ws.cell(row=row, column=10).value}
    return {}


def _route_lt_to_date(ws, date_str, lt_val):
    """Write Orlen LT (+ recomputed Δ) to the row matching the PDF's own date."""
    for row in range(5, ws.max_row + 1):
        c = ws.cell(row=row, column=1).value
        if c is None:
            continue
        ds = c.strftime('%Y-%m-%d') if hasattr(c, 'strftime') else str(c)[:10]
        if ds == date_str:
            ws.cell(row=row, column=6).value = lt_val
            ws.cell(row=row, column=6).number_format = '0.000'
            pl_eur = ws.cell(row=row, column=5).value
            if pl_eur and lt_val:
                ws.cell(row=row, column=7).value = round(pl_eur - lt_val, 4)
                ws.cell(row=row, column=7).number_format = '+0.000;-0.000;"-"'
                ws.cell(row=row, column=8).value = round((pl_eur - lt_val) / lt_val, 4)
                ws.cell(row=row, column=8).number_format = '+0.0%;-0.0%;"-"'
            log("Orlen LT", f"Routed LT={lt_val} → {date_str} (row {row})")
            return True
    log("Orlen LT", f"No existing row for {date_str}; LT not routed", "WARN")
    return False


# ═══════════════════════════════════════
# EXCEL WRITER
# ═══════════════════════════════════════
def update_excel(fx=None, orlen_pl=None, orlen_lt=None, elvis_de=None, bsh_se=None, eu_bulletin=None, _failures=None):
    global SUSPECTS
    SUSPECTS = []
    if _failures is None:
        _failures = {}
    if not EXCEL_PATH.exists():
        log("Excel", f"Not found: {EXCEL_PATH}", "ERROR")
        return False
    wb = load_workbook(str(EXCEL_PATH))

    # ELVIS DE: Tankerkönig value if present, else EC bulletin Germany diesel
    de_val = elvis_de["price_eur_l"] if elvis_de else None
    if de_val is None and eu_bulletin and eu_bulletin.get("DE"):
        de_val = eu_bulletin["DE"]

    if "Daily Tracker" in wb.sheetnames:
        ws = wb["Daily Tracker"]
        prev = _prev_daily_values(ws)
        existing_row = date_exists_in_daily(ws, TODAY)
        if existing_row:
            row = existing_row
            log("Excel", f"Updating existing row {row} for {TODAY_STR}")
        else:
            ws.insert_rows(5)
            row = 5
            log("Excel", f"Inserting new row {row} for {TODAY_STR}")

        ifont = Font(name=cfg.EXCEL_FONT_FAMILY, size=10, color="1D4ED8")
        ifill = PatternFill("solid", fgColor="EFF6FF")
        dfont = Font(name=cfg.EXCEL_FONT_FAMILY, size=10, color="1F2937")
        brd = Border(left=Side("thin", color="D1D5DB"), right=Side("thin", color="D1D5DB"),
                     top=Side("thin", color="D1D5DB"), bottom=Side("thin", color="D1D5DB"))

        def wc(col, val, fmt='General', font=dfont, fill=None):
            c = ws.cell(row=row, column=col)
            if val is not None or c.value is None:
                c.value = val
            c.number_format = fmt
            c.font = font
            c.border = brd
            c.alignment = Alignment(horizontal="right" if col > 2 else "center", vertical="center")
            if fill:
                c.fill = fill

        plnm3 = orlen_pl["price_pln_m3"] if orlen_pl else None
        lt_val = orlen_lt["price_eur_l"] if orlen_lt else None
        lt_date = orlen_lt.get("pdf_date") if orlen_lt else None
        # LT is written to today's row only if the PDF is actually today's;
        # otherwise it is routed to its real date row (see end of block).
        lt_for_today = lt_val if (lt_val is not None and (lt_date is None or lt_date == TODAY_STR)) else None
        se_sek = bsh_se["price_sek_l"] if bsh_se else None

        # ── Outlier guard (flag, don't drop) ──
        for name, new, key in [("PL", plnm3, "PL"), ("LT", lt_for_today, "LT"),
                               ("DE", de_val, "DE"), ("SE", se_sek, "SE")]:
            if new is not None and not validate_price_change(name, new, prev.get(key)):
                SUSPECTS.append(key)
        if fx and not validate_fx(fx.get("PLN_EUR"), fx.get("SEK_EUR")):
            SUSPECTS.append("FX")

        wc(1, TODAY, 'YYYY-MM-DD', Font(name=cfg.EXCEL_FONT_FAMILY, size=10, bold=True, color="1F2937"))
        wc(2, calendar.day_abbr[WDAY], font=Font(name=cfg.EXCEL_FONT_FAMILY, size=9, color="6B7280"))
        wc(3, plnm3, '#,##0.00', ifont, ifill)
        wc(4, fx["PLN_EUR"] if fx else None, '0.0000', ifont, ifill)
        # Polish wholesale is NETTO; add 23% VAT for parity with LT (su PVM)
        PL_PVM = 1.23
        pl_eur_l = round(plnm3 * PL_PVM / fx["PLN_EUR"] / 1000, 4) if (plnm3 and fx and fx.get("PLN_EUR")) else None
        wc(5, pl_eur_l, '0.000')
        wc(6, lt_for_today, '0.000', ifont, ifill)
        delta = round(pl_eur_l - lt_for_today, 4) if (pl_eur_l and lt_for_today) else None
        wc(7, delta, '+0.000;-0.000;"-"')
        delta_pct = round(delta / lt_for_today, 4) if (delta is not None and lt_for_today) else None
        wc(8, delta_pct, '+0.0%;-0.0%;"-"')
        wc(9, de_val, '0.000', ifont, ifill)
        wc(10, se_sek, '0.00', ifont, ifill)
        wc(11, fx["SEK_EUR"] if fx else None, '0.0000', ifont, ifill)
        se_eur = round(se_sek / fx["SEK_EUR"], 4) if (se_sek and fx and fx.get("SEK_EUR")) else None
        wc(12, se_eur, '0.000')

        ok = [k for k, v in {"FX": fx, "PL": orlen_pl, "LT": orlen_lt,
                             "DE": de_val, "SE": bsh_se, "EU": eu_bulletin}.items() if v]
        notes = f"Auto: {','.join(ok)}"
        if orlen_lt and orlen_lt.get("diesel_class"):
            notes += f"|CLS:{orlen_lt['diesel_class']}"
        if lt_val is not None and lt_date and lt_date != TODAY_STR:
            notes += f"|LT→{lt_date}"  # stale PDF: value routed to its real date
        if elvis_de and elvis_de.get("source"):
            notes += f"|DE:{elvis_de['source']}"
        elif de_val is not None:
            notes += "|DE:ec_bulletin"
        if _failures:
            notes += "|MISS:" + ','.join(f"{k}={v}" for k, v in _failures.items())
        if SUSPECTS:
            notes += "|SUSP:" + ','.join(SUSPECTS)
        wc(13, notes, font=Font(name=cfg.EXCEL_FONT_FAMILY, size=9, color="6B7280"))
        wc(14, "Auto", font=Font(name=cfg.EXCEL_FONT_FAMILY, size=8, color="9CA3AF"))
        log("Excel", f"Row {row}: {TODAY_STR} [{','.join(ok)}]" + (f" SUSPECT={SUSPECTS}" if SUSPECTS else ""))
        # Route a stale (older) Orlen LT PDF to its own date row, not today's
        if lt_val is not None and lt_date and lt_date != TODAY_STR:
            _route_lt_to_date(ws, lt_date, lt_val)

    if eu_bulletin and "Weekly Oil Bulletin" in wb.sheetnames:
        ws_w = wb["Weekly Oil Bulletin"]
        ec_date_str = eu_bulletin.get("_date")
        if ec_date_str:
            week_date = datetime.strptime(ec_date_str, "%Y-%m-%d")
        else:
            week_date = TODAY - timedelta(days=WDAY)
            ec_date_str = week_date.strftime("%Y-%m-%d")
        existing = date_exists_in_weekly(ws_w, ec_date_str)
        if existing:
            w_row = existing
        else:
            ws_w.insert_rows(4)
            w_row = 4
        ws_w.cell(row=w_row, column=1).value = week_date
        ws_w.cell(row=w_row, column=1).number_format = 'YYYY-MM-DD'
        ws_w.cell(row=w_row, column=1).font = Font(name=cfg.EXCEL_FONT_FAMILY, size=10, bold=True, color="1F2937")
        for k, col in cfg.EU_WEEKLY_COLUMNS.items():
            val = eu_bulletin.get(k)
            if val is not None:
                c = ws_w.cell(row=w_row, column=col)
                c.value = val
                c.number_format = '0.000'
                c.font = Font(name=cfg.EXCEL_FONT_FAMILY, size=10, color="1D4ED8")
        lt_v, eu_v = eu_bulletin.get("LT"), eu_bulletin.get("EU_AVG")
        ws_w.cell(row=w_row, column=9).value = round((lt_v - eu_v) / eu_v, 4) if (lt_v and eu_v) else None
        ws_w.cell(row=w_row, column=9).number_format = '+0.0%;-0.0%;"-"'

    wb.save(str(EXCEL_PATH))
    log("Excel", f"Saved: {EXCEL_PATH}")
    return True


# ═══════════════════════════════════════
# MAIN
# ═══════════════════════════════════════
def main():
    print(f"\n{'='*60}\n FUEL PRICE TRACKER v7 — {TODAY_STR}\n{'='*60}\n")
    print(f"  Diesel season classes: {get_season_classes()} (month={TODAY.month})\n")

    R, FAIL = {}, {}
    print("── FX Rates ──"); R["fx"] = fetch_fx()
    if not R["fx"]:
        FAIL["FX"] = "TIMEOUT"
    if WDAY < 5:
        print("\n── Orlen PL ──"); R["orlen_pl"] = fetch_orlen_pl()
        if not R["orlen_pl"]:
            FAIL["PL"] = "NO_DATA"
        print("\n── Orlen LT (PDF) ──"); R["orlen_lt"] = fetch_orlen_lt()
        if not R["orlen_lt"]:
            FAIL["LT"] = "NO_PDF"
        print("\n── Elvis DE ──"); R["elvis_de"] = fetch_elvis_de()
        # No FAIL for DE here: EC bulletin DE is the documented fallback
        print("\n── BSH/ST1 SE ──"); R["bsh_se"] = fetch_bsh_se()
        if not R["bsh_se"]:
            FAIL["SE"] = "NO_DATA"
    else:
        print("\n── Weekend ──")
        for k in ["orlen_pl", "orlen_lt", "elvis_de", "bsh_se"]:
            R[k] = None
        FAIL = {"PL": "WEEKEND", "LT": "WEEKEND", "DE": "WEEKEND", "SE": "WEEKEND"}
    print("\n── EU Oil Bulletin (EC XLSX) ──"); R["eu_bulletin"] = fetch_eu_bulletin()
    if not R["eu_bulletin"]:
        FAIL["EU"] = "NO_DATA"

    print(f"\n{'─'*60}")
    ok = sum(1 for v in R.values() if v is not None)
    print(f"RESULTS: {ok}/{len(R)}")
    for k, v in R.items():
        print(f"  {'✅' if v else '❌'} {k}: {v if v else 'FAILED'}")
    if FAIL:
        print(f"  Failures: {FAIL}")
    print(f"{'─'*60}\n── Updating Excel ──")
    success = update_excel(**R, _failures=FAIL)

    json_data = {
        "date": TODAY_STR, "version": "v7",
        "sources_ok": ok, "sources_total": len(R),
        "results": {k: bool(v) for k, v in R.items()},
        "data": {k: v for k, v in R.items() if v},
        "failures": FAIL,
        "suspects": SUSPECTS,
    }
    if R.get("orlen_lt") and R["orlen_lt"].get("diesel_class"):
        json_data["diesel_class"] = R["orlen_lt"]["diesel_class"]
    cfg.JSON_PATH.write_text(json.dumps(json_data, indent=2, default=str))

    print(f"\n{'✅ Done!' if success else '❌ Failed!'}")
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()

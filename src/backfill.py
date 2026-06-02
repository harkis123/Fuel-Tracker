"""
Backfill / repair fuel_tracker.xlsx — run once:  python src/backfill.py

Re-derives the Orlen LT column CORRECTLY for every date already in the sheet by
fetching each date's OWN PDF and reading the 'Pardavimo kaina su PVM' column
(same logic as the live scraper). This fixes the historical off-by-one where
each row showed the previous business day's price. Also gap-fills FX.

Run in your own environment (GitHub Actions "backfill" input = true, or locally).
"""
import io
import requests
from datetime import datetime
from openpyxl import load_workbook

import config as cfg
from scraper import parse_orlen_lt_pdf, _orlen_lt_url_for, H, EXCEL_PATH


def log(m):
    print(f"  {m}")


def repair_orlen_lt(ws):
    """For each date row, fetch that date's own PDF and set the correct su-PVM price."""
    fixed = same = missing = 0
    for r in range(5, ws.max_row + 1):
        dt = ws.cell(row=r, column=1).value
        if dt is None:
            continue
        ds = dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt)[:10]
        try:
            d = datetime.strptime(ds, '%Y-%m-%d')
        except ValueError:
            continue
        url = _orlen_lt_url_for(d)
        try:
            resp = requests.get(url, headers=H, timeout=cfg.REQUEST_TIMEOUT)
            if resp.status_code != 200 or len(resp.content) < 500:
                missing += 1
                continue
            res = parse_orlen_lt_pdf(resp.content)
        except Exception as e:
            log(f"{ds}: fetch/parse error {e}")
            missing += 1
            continue
        if not res:
            missing += 1
            continue
        # Only trust a PDF whose own validity date matches this row's date
        if res.get("pdf_date") and res["pdf_date"] != ds:
            log(f"{ds}: PDF date {res['pdf_date']} != row date — skipping")
            missing += 1
            continue
        new = res["price_eur_l"]
        old = ws.cell(row=r, column=6).value
        ws.cell(row=r, column=6).value = new
        ws.cell(row=r, column=6).number_format = '0.000'
        pl = ws.cell(row=r, column=5).value
        if pl and new:
            ws.cell(row=r, column=7).value = round(pl - new, 4)
            ws.cell(row=r, column=7).number_format = '+0.000;-0.000;"-"'
            ws.cell(row=r, column=8).value = round((pl - new) / new, 4)
            ws.cell(row=r, column=8).number_format = '+0.0%;-0.0%;"-"'
        if old != new:
            fixed += 1
            log(f"{ds}: LT {old} → {new}")
        else:
            same += 1
    return fixed, same, missing


def gapfill_fx(ws):
    """Fill any missing PLN/EUR (col4) or SEK/EUR (col11) from frankfurter history."""
    dates = []
    for r in range(5, ws.max_row + 1):
        dt = ws.cell(row=r, column=1).value
        if dt is None:
            continue
        ds = dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt)[:10]
        dates.append(ds)
    if not dates:
        return 0
    lo, hi = min(dates), max(dates)
    try:
        r = requests.get(f"https://api.frankfurter.app/{lo}..{hi}?from=EUR&to=PLN,SEK",
                         timeout=cfg.FX_TIMEOUT)
        fx = r.json().get("rates", {})
    except Exception as e:
        log(f"FX history failed: {e}")
        return 0
    n = 0
    for r in range(5, ws.max_row + 1):
        dt = ws.cell(row=r, column=1).value
        if dt is None:
            continue
        ds = dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt)[:10]
        day = fx.get(ds, {})
        if ws.cell(row=r, column=4).value is None and day.get("PLN"):
            ws.cell(row=r, column=4).value = day["PLN"]; n += 1
        if ws.cell(row=r, column=11).value is None and day.get("SEK"):
            ws.cell(row=r, column=11).value = day["SEK"]; n += 1
    return n


def main():
    print("=" * 60)
    print("  BACKFILL / REPAIR — Orlen LT per-date + FX gaps")
    print("=" * 60)
    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found")
        return
    wb = load_workbook(str(EXCEL_PATH))
    ws = wb["Daily Tracker"]

    print("\n[1/2] Repairing Orlen LT (own-date PDF, su PVM)...")
    fixed, same, missing = repair_orlen_lt(ws)
    log(f"LT: fixed={fixed}, already-correct={same}, no-PDF={missing}")

    print("\n[2/2] Gap-filling FX...")
    n = gapfill_fx(ws)
    log(f"FX cells filled: {n}")

    wb.save(str(EXCEL_PATH))
    print(f"\n  DONE — saved {EXCEL_PATH}")


if __name__ == "__main__":
    main()

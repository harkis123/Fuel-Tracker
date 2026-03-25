"""
Fuel Price Tracker v6
EC XLSX direct download, date from EC file, duplicate fix, DE from EC
"""
import requests, json, re, os, sys, io, calendar
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "fuel_tracker.xlsx"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36"
H = {"User-Agent": UA, "Accept": "text/html,application/xhtml+xml,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.9"}
TODAY = datetime.now()
TODAY_STR = TODAY.strftime("%Y-%m-%d")
WDAY = TODAY.weekday()

def log(s, m, l="INFO"): print(f"[{l}] {s}: {m}")

def clean_num(s):
    """Parse numbers like '1 756.06', '6 192', '1,234.56', '1234,56'"""
    if s is None: return None
    s = str(s).strip()
    # Remove currency symbols and whitespace-like chars
    s = re.sub(r'[€$£\xa0]', '', s)
    # Handle "1 756.06" — spaces as thousand separators, dot as decimal
    if re.match(r'^\d[\d ]+\.\d+$', s):
        s = s.replace(' ', '')
        return float(s)
    # Handle "1 756,06" — spaces as thousand sep, comma as decimal
    if re.match(r'^\d[\d ]+,\d+$', s):
        s = s.replace(' ', '').replace(',', '.')
        return float(s)
    # Handle "1,234.56"
    if ',' in s and '.' in s:
        s = s.replace(',', '')
        return float(s)
    # Handle "1234,56"
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    # Remove remaining spaces
    s = s.replace(' ', '')
    try: return float(s)
    except: return None

# ═══════════════════════════════════════
# 1. FX RATES — with multiple fallback APIs
# ═══════════════════════════════════════
def fetch_fx():
    import time
    
    # Source 1: frankfurter.app (ECB data)
    try:
        r = requests.get("https://api.frankfurter.app/latest?from=EUR&to=PLN,SEK", timeout=30)
        r.raise_for_status()
        d = r.json().get("rates", {})
        if d.get("PLN") and d.get("SEK"):
            log("FX", f"frankfurter.app: PLN={d['PLN']}, SEK={d['SEK']}")
            return {"PLN_EUR": d["PLN"], "SEK_EUR": d["SEK"]}
    except Exception as e:
        log("FX", f"frankfurter.app failed: {e}", "WARN")
    
    time.sleep(2)
    
    # Source 2: exchangerate.host (free, no key needed)
    try:
        r = requests.get("https://api.exchangerate.host/latest?base=EUR&symbols=PLN,SEK", timeout=30)
        r.raise_for_status()
        d = r.json().get("rates", {})
        if d.get("PLN") and d.get("SEK"):
            log("FX", f"exchangerate.host: PLN={d['PLN']}, SEK={d['SEK']}")
            return {"PLN_EUR": d["PLN"], "SEK_EUR": d["SEK"]}
    except Exception as e:
        log("FX", f"exchangerate.host failed: {e}", "WARN")
    
    time.sleep(2)
    
    # Source 3: ECB direct XML
    try:
        r = requests.get("https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml", timeout=30)
        r.raise_for_status()
        pln = re.search(r"currency='PLN'\s+rate='([\d.]+)'", r.text)
        sek = re.search(r"currency='SEK'\s+rate='([\d.]+)'", r.text)
        if pln and sek:
            pln_v, sek_v = float(pln.group(1)), float(sek.group(1))
            log("FX", f"ECB XML: PLN={pln_v}, SEK={sek_v}")
            return {"PLN_EUR": pln_v, "SEK_EUR": sek_v}
    except Exception as e:
        log("FX", f"ECB XML failed: {e}", "WARN")
    
    log("FX", "All 3 sources failed!", "ERROR")
    return None

# ═══════════════════════════════════════
# 2. ORLEN PL — via petrodom.pl
# ═══════════════════════════════════════
def fetch_orlen_pl():
    try:
        url = "https://www.petrodom.pl/en/current-wholesale-fuel-prices-provided-by-pkn-orlen/"
        r = requests.get(url, headers=H, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        
        # Method 1: Table parsing
        tables = soup.find_all("table")
        log("Orlen PL", f"Found {len(tables)} tables")
        for table in tables:
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
                for i, cell in enumerate(cells):
                    if "ekodiesel" in cell.lower() and "arktyczny" not in cell.lower() and "grzewczy" not in cell.lower():
                        log("Orlen PL", f"Found Ekodiesel cell: {repr(cell)}")
                        for j in range(i+1, min(i+3, len(cells))):
                            raw = cells[j]
                            log("Orlen PL", f"  Price cell: {repr(raw)}")
                            price = clean_num(raw)
                            if price and 3000 < price < 10000:
                                log("Orlen PL", f"Ekodiesel = {price} PLN/m³")
                                return {"price_pln_m3": price}
        
        # Method 2: Text fallback — search full page text
        text = soup.get_text(" ", strip=True)
        log("Orlen PL", f"Table parsing failed, trying text search...")
        m = re.search(r'[Ee]kodiesel[^0-9]{0,60}?(\d[\d\s\xa0]*\d)', text)
        if m:
            price = clean_num(m.group(1))
            if price and 3000 < price < 10000:
                log("Orlen PL", f"Ekodiesel (text fallback) = {price} PLN/m³")
                return {"price_pln_m3": price}
        
        log("Orlen PL", "Not found in tables or text", "WARN")
        return None
    except Exception as e:
        log("Orlen PL", str(e), "ERROR"); return None

# ═══════════════════════════════════════
# 3. ORLEN LT — PDF: pardavimo kaina su PVM
# ═══════════════════════════════════════
def fetch_orlen_lt():
    try:
        list_url = "https://www.orlenlietuva.lt/lt/wholesale/_layouts/f2hPriceTable/default.aspx"
        r = requests.get(list_url, headers=H, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        pdf_links = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if ".pdf" in href.lower() and "kainos" in href.lower():
                if not href.startswith("http"):
                    href = "https://www.orlenlietuva.lt" + href
                pdf_links.append(href)
        if not pdf_links:
            for days_back in range(7):
                d = TODAY - timedelta(days=days_back)
                pdf_links.append(f"https://www.orlenlietuva.lt/LT/Wholesale/Prices/Kainos {d.strftime('%Y %m %d')} realizacija internet.pdf")
        for pdf_url in pdf_links[:5]:
            try:
                log("Orlen LT", f"Trying: {pdf_url.split('/')[-1]}")
                r2 = requests.get(pdf_url, headers=H, timeout=15)
                if r2.status_code == 200 and len(r2.content) > 500:
                    price = parse_orlen_lt_pdf(r2.content)
                    if price: return price
            except: continue
        log("Orlen LT", "No PDF parsed", "WARN"); return None
    except Exception as e:
        log("Orlen LT", str(e), "ERROR"); return None

def parse_orlen_lt_pdf(pdf_bytes):
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ""
            for line in text.split('\n'):
                if 'Dyzelinas E kl. su RRME' not in line:
                    continue
                # Extract numbers like "897.69" and "1 756.06"
                nums = re.findall(r'(\d[\d ]*\.\d{2})', line)
                cleaned = [float(n.replace(' ', '')) for n in nums]
                if cleaned:
                    # LAST number = Pardavimo kaina su PVM (EUR/1000l)
                    selling_price = cleaned[-1]
                    if 1000 < selling_price < 2500:
                        eur_l = round(selling_price / 1000, 4)
                        log("Orlen LT", f"Juodeikiai: {selling_price} EUR/1000l = {eur_l} EUR/l (su PVM)")
                        return {"price_eur_l": eur_l}
                break
    except ImportError:
        log("Orlen LT", "pdfplumber not installed", "WARN")
    except Exception as e:
        log("Orlen LT", f"PDF error: {e}", "WARN")
    return None

# ═══════════════════════════════════════
# 4. ELVIS DE — DIESEL from fuel-prices.eu
# ═══════════════════════════════════════
def fetch_elvis_de():
    try:
        r = requests.get("https://www.fuel-prices.eu/cheapest/", headers=H, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        tables = soup.find_all("table")
        for table in tables:
            if "diesel" not in table.get_text(" ", strip=True).lower(): continue
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
                row_text = " ".join(cells)
                if "germany" in row_text.lower() or " DE" in row_text:
                    for cell in cells:
                        m = re.search(r'€?(\d\.\d{3})', cell)
                        if m:
                            price = float(m.group(1))
                            if 0.8 < price < 3.5:
                                log("Elvis DE", f"Germany Diesel = {price} EUR/l")
                                return {"price_eur_l": price}
        log("Elvis DE", "Not found", "WARN"); return None
    except Exception as e:
        log("Elvis DE", str(e), "ERROR"); return None

# ═══════════════════════════════════════
# 5. BSH / ST1 SE
# ═══════════════════════════════════════
def fetch_bsh_se():
    try:
        r = requests.get("https://st1.se/foretag/listpris", headers=H, timeout=20)
        r.raise_for_status()
        text = BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True)
        for pat in [r'[Dd]iesel\s*(?:MK[123])?\s*[^0-9]{0,30}?(\d{1,2}[.,]\d{2})', r'(\d{1,2}[.,]\d{2})\s*(?:kr|SEK)[^0-9]{0,20}[Dd]iesel']:
            m = re.search(pat, text)
            if m:
                p = float(m.group(1).replace(",", "."))
                if 10 < p < 35:
                    log("BSH SE", f"Diesel = {p} SEK/l")
                    return {"price_sek_l": p}
        log("BSH SE", "Not found", "WARN"); return None
    except Exception as e:
        log("BSH SE", str(e), "ERROR"); return None

# ═══════════════════════════════════════
# 6. EU BULLETIN — direct EC XLSX download
# ═══════════════════════════════════════
def fetch_eu_bulletin():
    """Download EC Weekly Oil Bulletin XLSX — take DIESEL column (Gas oil automobile), not Euro-super 95"""
    try:
        url = "https://energy.ec.europa.eu/document/download/264c2d0f-f161-4ea3-a777-78faae59bea0_en?filename=Weekly%20Oil%20Bulletin%20Weekly%20prices%20with%20Taxes%20-%202024-02-19.xlsx"
        log("EU Bulletin", "Downloading EC XLSX...")
        r = requests.get(url, headers=H, timeout=30)
        if r.status_code != 200:
            log("EU Bulletin", f"EC HTTP {r.status_code}", "WARN")
            return fetch_eu_bulletin_fallback()

        from openpyxl import load_workbook as lwb
        wb = lwb(io.BytesIO(r.content), data_only=True)

        countries = {"LT": None, "LV": None, "EE": None, "DK": None, "SE": None, "FI": None}
        cc_names = {"Lithuania": "LT", "Latvia": "LV", "Estonia": "EE",
                    "Denmark": "DK", "Sweden": "SE", "Finland": "FI"}
        eu_avg = None
        de_diesel = None
        ec_date = None

        ws = wb[wb.sheetnames[0]]  # first sheet
        log("EU Bulletin", f"Sheet: {ws.title} ({ws.max_row}r x {ws.max_column}c)")

        # Step 1: Find the DIESEL column by scanning headers
        # EC structure: col A=country, col B=Euro-super 95, col C=Gas oil (DIESEL), col D=Heating, ...
        # But let's find it dynamically in case it shifts
        diesel_col = 3  # default: column C
        for row in range(1, 6):
            for col in range(1, min(10, ws.max_column + 1)):
                val = str(ws.cell(row=row, column=col).value or "").lower()
                if "gas oil" in val or "diesel" in val or "gasoil" in val:
                    diesel_col = col
                    log("EU Bulletin", f"Diesel column found: {col} (row {row}: '{val[:50]}')")
                    break
            if diesel_col != 3: break

        log("EU Bulletin", f"Using diesel column: {diesel_col}")

        # Step 2: Find date in first few rows
        for row in range(1, 6):
            for col in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val is None: continue
                if hasattr(val, 'strftime'):
                    ec_date = val.strftime('%Y-%m-%d')
                    log("EU Bulletin", f"Date (datetime): {ec_date}")
                elif isinstance(val, str):
                    dm = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', val)
                    if dm:
                        d, m, y = dm.group(1), dm.group(2), dm.group(3)
                        ec_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                        log("EU Bulletin", f"Date (text): {ec_date}")

        # Step 3: Read country prices from DIESEL column only
        for row in range(1, ws.max_row + 1):
            cell0 = str(ws.cell(row=row, column=1).value or "").strip()
            if not cell0: continue

            # Match our target countries
            for cname, cc in cc_names.items():
                if cname.lower() in cell0.lower():
                    val = ws.cell(row=row, column=diesel_col).value
                    if val is not None:
                        try:
                            v = float(val)
                            if 0.5 < v < 3.5:
                                countries[cc] = round(v, 4)
                            elif 500 < v < 3500:
                                countries[cc] = round(v / 1000, 4)
                            log("EU Bulletin", f"  {cname}: {v} → {countries[cc]} EUR/l")
                        except: pass

            # Germany
            if "germany" in cell0.lower():
                val = ws.cell(row=row, column=diesel_col).value
                if val is not None:
                    try:
                        v = float(val)
                        if 0.5 < v < 3.5: de_diesel = round(v, 4)
                        elif 500 < v < 3500: de_diesel = round(v / 1000, 4)
                        log("EU Bulletin", f"  Germany: {v} → {de_diesel} EUR/l")
                    except: pass

            # EU weighted average — take EU27, skip Euro Area
            c0l = cell0.lower()
            if eu_avg is None and ("ce/ec" in c0l or "eur27" in c0l or "eu" in c0l) and ("average" in c0l or "weighted" in c0l or "moyenne" in c0l or "durchschnitt" in c0l):
                # Skip "Euro Area" — we want EU27
                if "euro area" in c0l or "eurozone" in c0l:
                    log("EU Bulletin", f"  Skipping Euro Area row: {cell0[:60]}")
                else:
                    val = ws.cell(row=row, column=diesel_col).value
                    if val is not None:
                        try:
                            v = float(val)
                            if 0.5 < v < 3.5: eu_avg = round(v, 4)
                            elif 500 < v < 3500: eu_avg = round(v / 1000, 4)
                            log("EU Bulletin", f"  EU27 avg: {v} → {eu_avg} EUR/l")
                        except: pass

        found = {k: v for k, v in countries.items() if v is not None}
        if found:
            log("EU Bulletin", f"DIESEL prices: {found}, avg={eu_avg}, DE={de_diesel}, date={ec_date}")
            result = {**countries, "EU_AVG": eu_avg}
            if de_diesel: result["DE"] = de_diesel
            if ec_date: result["_date"] = ec_date
            return result

        log("EU Bulletin", "No diesel data found in EC XLSX", "WARN")
        return fetch_eu_bulletin_fallback()

    except Exception as e:
        log("EU Bulletin", f"EC XLSX error: {e}", "WARN")
        return fetch_eu_bulletin_fallback()


def fetch_eu_bulletin_fallback():
    """Fallback: fuel-prices.eu"""
    try:
        r = requests.get("https://www.fuel-prices.eu/cheapest/", headers=H, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        countries = {"LT": None, "LV": None, "EE": None, "DK": None, "SE": None, "FI": None}
        cc_map = {"lithuania":"LT","latvia":"LV","estonia":"EE","denmark":"DK","sweden":"SE","finland":"FI"}
        eu_avg = None
        for table in soup.find_all("table"):
            if "diesel" not in table.get_text(" ", strip=True).lower(): continue
            for row in table.find_all("tr"):
                cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
                if len(cells) < 2: continue
                rt = " ".join(cells[:2]).lower()
                for name, cc in cc_map.items():
                    if name in rt:
                        for cell in cells:
                            m = re.search(r'€?(\d\.\d{3})', cell)
                            if m:
                                val = float(m.group(1))
                                if 0.5 < val < 3.5: countries[cc] = val; break
        text = soup.get_text(" ", strip=True)
        m = re.search(r'€(\d\.\d{3})/L\s+for\s+diesel', text)
        if m: eu_avg = float(m.group(1))
        # Try to find date
        dm = re.search(r'(\w+\s+\d{1,2},?\s+\d{4})', text)
        found = {k:v for k,v in countries.items() if v}
        if found:
            log("EU Bulletin fallback", f"{found}, avg={eu_avg}")
            return {**countries, "EU_AVG": eu_avg}
        return None
    except Exception as e:
        log("EU Bulletin fallback", str(e), "ERROR"); return None

# ═══════════════════════════════════════
# EXCEL: check if date already exists
# ═══════════════════════════════════════
def date_exists_in_daily(ws, target_date):
    """Check if today's date already has data in Daily Tracker"""
    for row in range(5, min(15, ws.max_row + 1)):
        cell = ws.cell(row=row, column=1).value
        if cell is None: continue
        if hasattr(cell, 'date'):
            if cell.date() == target_date.date(): return row
        elif hasattr(cell, 'strftime'):
            if cell.strftime('%Y-%m-%d') == target_date.strftime('%Y-%m-%d'): return row
        elif isinstance(cell, str) and cell[:10] == target_date.strftime('%Y-%m-%d'):
            return row
    return None

def date_exists_in_weekly(ws, date_str):
    """Check if this week's data already exists. date_str = 'YYYY-MM-DD'"""
    for row in range(4, min(ws.max_row + 1, 100)):
        cell = ws.cell(row=row, column=1).value
        if cell is None: continue
        if hasattr(cell, 'strftime'):
            if cell.strftime('%Y-%m-%d') == date_str: return row
        elif hasattr(cell, 'date'):
            if cell.date().strftime('%Y-%m-%d') == date_str: return row
        elif isinstance(cell, str) and cell[:10] == date_str:
            return row
    return None

# ═══════════════════════════════════════
# EXCEL WRITER
# ═══════════════════════════════════════
def update_excel(fx=None, orlen_pl=None, orlen_lt=None, elvis_de=None, bsh_se=None, eu_bulletin=None, _failures=None):
    if _failures is None: _failures = {}
    if not EXCEL_PATH.exists():
        log("Excel", f"Not found: {EXCEL_PATH}", "ERROR"); return False
    wb = load_workbook(str(EXCEL_PATH))

    if "Daily Tracker" in wb.sheetnames:
        ws = wb["Daily Tracker"]
        
        # Check if today already exists — update instead of insert
        existing_row = date_exists_in_daily(ws, TODAY)
        if existing_row:
            row = existing_row
            log("Excel", f"Updating existing row {row} for {TODAY_STR}")
        else:
            ws.insert_rows(5)
            row = 5
            log("Excel", f"Inserting new row {row} for {TODAY_STR}")

        ifont = Font(name="Aptos", size=10, color="1D4ED8")
        ifill = PatternFill("solid", fgColor="EFF6FF")
        dfont = Font(name="Aptos", size=10, color="1F2937")
        brd = Border(left=Side("thin",color="D1D5DB"),right=Side("thin",color="D1D5DB"),top=Side("thin",color="D1D5DB"),bottom=Side("thin",color="D1D5DB"))

        def wc(col, val, fmt='General', font=dfont, fill=None):
            c = ws.cell(row=row, column=col)
            # Only overwrite if we have new data, keep existing if None
            if val is not None or c.value is None:
                c.value = val
            c.number_format = fmt; c.font = font; c.border = brd
            c.alignment = Alignment(horizontal="right" if col > 2 else "center", vertical="center")
            if fill: c.fill = fill

        wc(1, TODAY, 'YYYY-MM-DD', Font(name="Aptos",size=10,bold=True,color="1F2937"))
        wc(2, calendar.day_abbr[WDAY], font=Font(name="Aptos",size=9,color="6B7280"))
        wc(3, orlen_pl["price_pln_m3"] if orlen_pl else None, '#,##0.00', ifont, ifill)
        wc(4, fx["PLN_EUR"] if fx else None, '0.0000', ifont, ifill)
        
        pl_eur_l = round(orlen_pl["price_pln_m3"] / fx["PLN_EUR"] / 1000, 4) if (orlen_pl and fx and fx.get("PLN_EUR")) else None
        wc(5, pl_eur_l, '0.000')
        
        lt_val = orlen_lt["price_eur_l"] if orlen_lt else None
        wc(6, lt_val, '0.000', ifont, ifill)
        
        delta = round(pl_eur_l - lt_val, 4) if (pl_eur_l and lt_val) else None
        wc(7, delta, '+0.000;-0.000;"-"')
        delta_pct = round(delta / lt_val, 4) if (delta is not None and lt_val) else None
        wc(8, delta_pct, '+0.0%;-0.0%;"-"')
        
        wc(9, elvis_de["price_eur_l"] if elvis_de else None, '0.000', ifont, ifill)
        # If EC has Germany diesel, prefer it over fuel-prices.eu
        if eu_bulletin and eu_bulletin.get("DE"):
            wc(9, eu_bulletin["DE"], '0.000', ifont, ifill)

        wc(10, bsh_se["price_sek_l"] if bsh_se else None, '0.00', ifont, ifill)
        wc(11, fx["SEK_EUR"] if fx else None, '0.0000', ifont, ifill)
        
        se_eur = round(bsh_se["price_sek_l"] / fx["SEK_EUR"], 4) if (bsh_se and fx and fx.get("SEK_EUR")) else None
        wc(12, se_eur, '0.000')
        
        ok = [k for k,v in {"FX":fx,"PL":orlen_pl,"LT":orlen_lt,"DE":(elvis_de or (eu_bulletin and eu_bulletin.get("DE"))),"SE":bsh_se,"EU":eu_bulletin}.items() if v]
        notes = f"Auto: {','.join(ok)}"
        if _failures:
            miss = ','.join([f"{k}={v}" for k,v in _failures.items()])
            notes += f"|MISS:{miss}"
        wc(13, notes, font=Font(name="Aptos",size=9,color="6B7280"))
        wc(14, "Auto", font=Font(name="Aptos",size=8,color="9CA3AF"))
        log("Excel", f"Row {row}: {TODAY_STR} [{','.join(ok)}]")

    # Weekly — use actual EC date, not calculated Monday
    if eu_bulletin and "Weekly Oil Bulletin" in wb.sheetnames:
        ws_w = wb["Weekly Oil Bulletin"]
        
        # Use date from EC file if available, otherwise calculate Monday
        ec_date_str = eu_bulletin.get("_date")
        if ec_date_str:
            week_date = datetime.strptime(ec_date_str, "%Y-%m-%d")
            log("Excel", f"Using EC date: {ec_date_str}")
        else:
            week_date = TODAY - timedelta(days=WDAY)
            ec_date_str = week_date.strftime("%Y-%m-%d")
            log("Excel", f"Using calculated Monday: {ec_date_str}")
        
        existing = date_exists_in_weekly(ws_w, ec_date_str)
        if existing:
            w_row = existing
            log("Excel", f"Updating weekly row {w_row} for {ec_date_str}")
        else:
            ws_w.insert_rows(4)
            w_row = 4
            log("Excel", f"Inserting weekly row {w_row} for {ec_date_str}")
        
        ws_w.cell(row=w_row,column=1).value = week_date
        ws_w.cell(row=w_row,column=1).number_format = 'YYYY-MM-DD'
        ws_w.cell(row=w_row,column=1).font = Font(name="Aptos",size=10,bold=True,color="1F2937")
        for k,col in {"LT":2,"LV":3,"EE":4,"DK":5,"SE":6,"FI":7,"EU_AVG":8}.items():
            val = eu_bulletin.get(k)
            if val is not None:
                c = ws_w.cell(row=w_row,column=col)
                c.value = val; c.number_format = '0.000'
                c.font = Font(name="Aptos",size=10,color="1D4ED8")
        lt_v = eu_bulletin.get("LT"); eu_v = eu_bulletin.get("EU_AVG")
        ws_w.cell(row=w_row,column=9).value = round((lt_v-eu_v)/eu_v,4) if (lt_v and eu_v) else None
        ws_w.cell(row=w_row,column=9).number_format = '+0.0%;-0.0%;"-"'

    wb.save(str(EXCEL_PATH))
    log("Excel", f"Saved: {EXCEL_PATH}")
    return True

# ═══════════════════════════════════════
# MAIN
# ═══════════════════════════════════════
def main():
    print(f"\n{'='*60}\n  FUEL PRICE TRACKER v6 — {TODAY_STR}\n{'='*60}\n")
    R = {}
    FAIL = {}  # track why each source failed
    
    print("── FX Rates ──"); R["fx"] = fetch_fx()
    if not R["fx"]: FAIL["FX"] = "TIMEOUT"
    
    if WDAY < 5:
        print("\n── Orlen PL ──"); R["orlen_pl"] = fetch_orlen_pl()
        if not R["orlen_pl"]: FAIL["PL"] = "NO_DATA"
        
        print("\n── Orlen LT (PDF) ──"); R["orlen_lt"] = fetch_orlen_lt()
        if not R["orlen_lt"]: FAIL["LT"] = "NO_PDF"
        
        print("\n── Elvis DE (Diesel) ──"); R["elvis_de"] = fetch_elvis_de()
        if not R["elvis_de"]: FAIL["DE"] = "NO_DATA"
        
        print("\n── BSH/ST1 SE ──"); R["bsh_se"] = fetch_bsh_se()
        if not R["bsh_se"]: FAIL["SE"] = "NO_DATA"
    else:
        print("\n── Weekend ──")
        for k in ["orlen_pl","orlen_lt","elvis_de","bsh_se"]: R[k] = None
        FAIL = {"PL":"WEEKEND","LT":"WEEKEND","DE":"WEEKEND","SE":"WEEKEND"}
    
    print("\n── EU Oil Bulletin (EC XLSX) ──"); R["eu_bulletin"] = fetch_eu_bulletin()
    if not R["eu_bulletin"]: FAIL["EU"] = "NO_DATA"
    
    print(f"\n{'─'*60}")
    ok = sum(1 for v in R.values() if v is not None)
    print(f"RESULTS: {ok}/{len(R)}")
    for k,v in R.items(): print(f"  {'✅' if v else '❌'} {k}: {v if v else 'FAILED'}")
    if FAIL: print(f"  Failures: {FAIL}")
    print(f"{'─'*60}\n── Updating Excel ──")
    success = update_excel(**R, _failures=FAIL)
    Path(EXCEL_PATH.parent / "latest_results.json").write_text(
        json.dumps({"date":TODAY_STR,"results":{k:bool(v) for k,v in R.items()},"data":{k:v for k,v in R.items() if v},"failures":FAIL}, indent=2, default=str))
    print(f"\n{'✅ Done!' if success else '❌ Failed!'}")
    if not success: sys.exit(1)

if __name__ == "__main__":
    main()
